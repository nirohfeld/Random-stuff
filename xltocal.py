from __future__ import print_function
from apiclient.discovery import build
from httplib2 import Http
from oauth2client import file, client, tools
import openpyxl
import datetime

try:
    import argparse
    flags = argparse.ArgumentParser(parents=[tools.argparser]).parse_args()
except ImportError:
    flags = None

SCOPES = 'https://www.googleapis.com/auth/calendar'
store = file.Storage('storage.json')
creds = store.get()
if not creds or creds.invalid:
    flow = client.flow_from_clientsecrets('client_secret.json', SCOPES)
    creds = tools.run_flow(flow, store,flags)
CAL = build('calendar', 'v3', http=creds.authorize(Http()))

def addEvent(eventName,date): 
    EVENT = {
        'summary': eventName,
        'start':  {"date": date,},
        'end':    {"date": date,},
    }

    e = CAL.events().insert(calendarId='primary',
            sendNotifications=True, body=EVENT).execute()

    print('''*** %r event added:
        Start: %s
        End:   %s''' % (e['summary'].encode('utf-8'),
            e['start']['date'], e['end']['date']))

"""gets a column and the excel sheet and return the date in that column in a a formated string"""
def get_event_date(column,ws):
        date = 'N' #the dates are in the N column
        date = date + str(column)
        emtpy_string = ws['N41'].value #the value at a empty string
        if (ws[date].value == emtpy_string):
            return -1
        if (ws[date].value.month > 9):
            date_string = '2015-'+str(ws[date].value.month)+'-'+str(ws[date].value.day)
        else:
            date_string = '2016-0'+str(ws[date].value.month)+'-'+str(ws[date].value.day) #TD:catch when the year is changing
        return date_string

wb = openpyxl.load_workbook('tests.xlsx')
ws = wb.worksheets[0]
for x in range(100,131):
    date = 'K' #the events are in the K column
    date = date + str(x)
    emtpy_string = ws['N41'].value #the value at a empty string
    weekend_string = ws['K4'].value #the value at a weekend event
    if(ws[date].value == emtpy_string or ws[date].value == weekend_string):
        continue
    eventName = ws[date].value
    if(get_event_date(x,ws) != -1):
        eventDate = get_event_date(x,ws)
        addEvent(eventName,eventDate)
